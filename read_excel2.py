from openpyxl import load_workbook
import sys

try:
    wb = load_workbook(r'C:\Users\liu\Desktop\主账本_Liu.xlsx')
    print("工作表列表:", wb.sheetnames)

    ws = wb.active
    print(f"\n活动工作表: {ws.title}")

    # 读取前15行
    print("\n前15行数据:")
    for i, row in enumerate(ws.iter_rows(max_row=15, values_only=True), 1):
        print(f"{i}: {row}")

except Exception as e:
    print(f"错误: {e}", file=sys.stderr)
    import traceback
    traceback.print_exc()
